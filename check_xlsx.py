import openpyxl
import sys

wb = openpyxl.load_workbook(sys.argv[1])
print('Sheet names:', wb.sheetnames)
ws = wb.worksheets[0]
print('Active sheet name:', ws.title)
print('Max row:', ws.max_row)
print('Max col:', ws.max_column)
print('\nFirst 5 rows:')
for i in range(1, 6):
    row = [ws.cell(i, j).value for j in range(1, 4)]
    print(f'Row {i}: {row}')
