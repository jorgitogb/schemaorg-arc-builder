"""Inspect investigation Excel file."""
from openpyxl import load_workbook
import sys

if len(sys.argv) < 2:
    print("Usage: python inspect_investigation.py <excel_file>")
    sys.exit(1)

try:
    wb = load_workbook(sys.argv[1])
except FileNotFoundError:
    print(f"Error: File '{sys.argv[1]}' not found.")
    sys.exit(1)
except Exception as e:
    print(f"Error loading workbook: {e}")
    sys.exit(1)

ws = wb.active

if ws is None:
    print("Error: No active worksheet found in the workbook.")
    print(f"Available sheets: {wb.sheetnames}")
    sys.exit(1)

print("Investigation Excel Content:")
print("=" * 80)
for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True), 1):
    print(f"Row {i:2d}: {row}")
